"""
Automated Revenue Extraction ETL Pipeline

This script scans public-company annual report PDFs, extracts revenue-related tables,
detects reporting units, cleans numeric values, and writes structured results to Excel.

Privacy / security notes:
- No local machine paths are hardcoded.
- Input and output paths are provided through command-line arguments.
- Do not upload proprietary annual reports or confidential data to a public repository.
"""

from __future__ import annotations

import argparse
import logging
import re
import shutil
from pathlib import Path
from typing import Dict, List, Optional

import pdfplumber
from openpyxl import Workbook, load_workbook
from tqdm import tqdm


# Source-document keyword and unit patterns.
# These matching terms are intentionally kept in Chinese because the source filings are Chinese-language PDFs.
KEYWORD_TABLE = re.compile(r"(营业收入|营业总收入|營業總收入|營業收入|营业收)")
KEYWORD_OWNER = re.compile(r"(2021 年|主要会计)")
EXCLUDE_KEYWORDS = re.compile(r"(季度|研发投入)")
ADJ_BEFORE = "调整前"

SUMMARY_SHEET = "Summary"
COL_FILE_NAME = "File Name"
COL_TABLE_PAGE = "Keyword Table Page"
COL_UNIT_TEXT = "Unit Text"
COL_UNIT = "Unit"
COL_Y2023 = "2023"
COL_Y2022 = "2022"
COL_Y2021 = "2021"
COL_WAS_MODIFIED = "Modified"

VAL_NO_DATA = "No Data"
VAL_NO_UNIT = "No Unit"
VAL_NO = "No"
VAL_YES = "Yes"
VAL_ERROR = "Error"

SUMMARY_COLUMNS = [
    COL_FILE_NAME,
    COL_TABLE_PAGE,
    COL_UNIT_TEXT,
    COL_UNIT,
    COL_Y2023,
    COL_Y2022,
    COL_Y2021,
    COL_WAS_MODIFIED,
]


def extract_unit(text_all: str) -> str:
    """Extract the reporting unit from nearby text around the target table."""
    pattern1 = r"(?:营业收入金额|营业收入|营业总收入|经营业绩)\s*（([^）]*?)）"
    match1 = re.search(pattern1, text_all)
    if match1:
        return match1.group(1).strip()

    pattern2 = r"单位：(.*?元)"
    match2 = re.search(pattern2, text_all)
    if match2:
        return match2.group(1).strip()

    pattern3 = r"（[^）]*元[^）]*）"
    matches3 = re.findall(pattern3, text_all)
    for match in matches3:
        return match.strip("（）").strip()

    return VAL_NO_DATA


def clean_illegal_chars(text: object) -> str:
    """Remove characters that can break Excel writing."""
    return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", str(text))


def fix_thousands_separator(value: str) -> str:
    """
    Fix malformed thousands separators when PDF extraction drops trailing zeros.

    Example:
        "2,697,350,02" -> "2,697,350,020"
    """
    if "," not in value:
        return value

    if "." in value:
        int_part, dec_part = value.split(".", 1)
    else:
        int_part, dec_part = value, ""

    parts = int_part.split(",")
    if len(parts) > 1 and len(parts[-1]) < 3:
        parts[-1] = parts[-1].ljust(3, "0")

    int_part = ",".join(parts)
    return f"{int_part}.{dec_part}" if dec_part else int_part


def ensure_excel_exists(excel_path: Path) -> None:
    """Create an output workbook with a Summary sheet if it does not already exist."""
    if excel_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET
    ws.append(SUMMARY_COLUMNS)
    wb.save(excel_path)


def safe_sheet_name(file_stem: str, existing_names: set[str]) -> str:
    """Create a valid unique Excel sheet name from a file name."""
    cleaned = re.sub(r"[\[\]\*\?/\\:]", "_", file_stem)[:31] or "PDF"
    candidate = cleaned
    counter = 1

    while candidate in existing_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1

    existing_names.add(candidate)
    return candidate


def extract_numeric_values(values: List[str]) -> tuple[List[float], bool]:
    """Clean extracted row values and keep numeric values likely to be revenue figures."""
    numeric_pattern = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d{2})?$")
    cleaned_values: List[str] = []
    modified = False

    for value in values:
        value = value.strip()
        if numeric_pattern.fullmatch(value):
            cleaned_values.append(value)
        else:
            fixed = fix_thousands_separator(value)
            cleaned_values.append(fixed)
            if fixed != value:
                modified = True

    nums: List[float] = []
    for value in cleaned_values:
        if not numeric_pattern.fullmatch(value):
            continue
        try:
            parsed = float(value.replace(",", ""))
        except ValueError:
            continue
        if parsed > 1000:
            nums.append(parsed)

    return nums, modified


def process_pdf(pdf_path: Path, workbook, existing_sheet_names: set[str], start_page: int, max_pages: int) -> Dict[str, object]:
    """Process one PDF and return a summary row dictionary."""
    summary: Dict[str, object] = {
        COL_FILE_NAME: pdf_path.name,
        COL_TABLE_PAGE: "",
        COL_UNIT_TEXT: "",
        COL_UNIT: VAL_NO_UNIT,
        COL_Y2023: "",
        COL_Y2022: "",
        COL_Y2021: "",
        COL_WAS_MODIFIED: VAL_NO,
    }

    sheet_title = safe_sheet_name(pdf_path.stem, existing_sheet_names)
    worksheet = workbook.create_sheet(title=sheet_title)

    with pdfplumber.open(str(pdf_path)) as pdf:
        start_index = max(start_page - 1, 0)
        end_index = min(start_index + max_pages, len(pdf.pages))

        for page_index in range(start_index, end_index):
            page = pdf.pages[page_index]
            tables = page.find_tables(
                table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 1,
                    "join_tolerance": 1,
                    "min_words_vertical": 1,
                    "min_words_horizontal": 1,
                }
            )

            for table in tables:
                raw_table = table.extract()
                flat_text = "\n".join(",".join(map(str, row)) for row in raw_table)

                if not (
                    KEYWORD_TABLE.search(flat_text)
                    and KEYWORD_OWNER.search(flat_text)
                    and not EXCLUDE_KEYWORDS.search(flat_text)
                ):
                    continue

                for row_idx, row in enumerate(raw_table, start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        value = str(cell).replace("\n", "").strip()
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

                summary[COL_TABLE_PAGE] = page_index + 1

                nearby_text: List[str] = []
                for adjacent_index in (page_index - 1, page_index, page_index + 1):
                    if 0 <= adjacent_index < len(pdf.pages):
                        nearby_text.append(pdf.pages[adjacent_index].extract_text() or "")

                unit_text = "\n".join(nearby_text)
                cleaned_unit_text = clean_illegal_chars(unit_text)
                summary[COL_UNIT_TEXT] = cleaned_unit_text
                summary[COL_UNIT] = extract_unit(cleaned_unit_text)

                target_row_values: List[str] = []
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    if any(KEYWORD_TABLE.search(str(cell.value or "")) for cell in row):
                        target_row_values = [
                            str(cell.value or "").replace("\n", "").strip()
                            for cell in row
                            if cell.value not in (None, "")
                        ]
                        break

                nums, modified = extract_numeric_values(target_row_values)
                summary[COL_WAS_MODIFIED] = VAL_YES if modified else VAL_NO

                picks = [0, 1, 3] if ADJ_BEFORE in unit_text and len(nums) >= 4 else [0, 1, 2]
                for idx, year in enumerate([COL_Y2023, COL_Y2022, COL_Y2021]):
                    try:
                        summary[year] = f"{nums[picks[idx]]:,.0f}"
                    except (IndexError, TypeError):
                        summary[year] = ""

                return summary

    return summary


def run_pipeline(
    input_dir: Path,
    output_file: Path,
    processed_dir: Optional[Path],
    failed_dir: Optional[Path],
    move_files: bool,
    start_page: int,
    max_pages: int,
) -> None:
    """Run the extraction pipeline over all PDFs in the input directory."""
    input_dir = input_dir.expanduser().resolve()
    output_file = output_file.expanduser().resolve()

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    if processed_dir:
        processed_dir = processed_dir.expanduser().resolve()
        processed_dir.mkdir(parents=True, exist_ok=True)

    if failed_dir:
        failed_dir = failed_dir.expanduser().resolve()
        failed_dir.mkdir(parents=True, exist_ok=True)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    ensure_excel_exists(output_file)

    workbook = load_workbook(output_file)
    summary_sheet = workbook[SUMMARY_SHEET]
    existing_sheet_names = set(workbook.sheetnames)

    pdf_files = sorted(input_dir.glob("*.pdf"))

    for pdf_path in tqdm(pdf_files, desc="Processing PDFs"):
        found = False
        try:
            summary = process_pdf(
                pdf_path=pdf_path,
                workbook=workbook,
                existing_sheet_names=existing_sheet_names,
                start_page=start_page,
                max_pages=max_pages,
            )
            found = bool(summary.get(COL_TABLE_PAGE))
            summary_sheet.append([summary[column] for column in SUMMARY_COLUMNS])

            if move_files and processed_dir and failed_dir:
                target_dir = processed_dir if found else failed_dir
                shutil.move(str(pdf_path), str(target_dir / pdf_path.name))

        except Exception as exc:
            logging.exception("Error processing %s", pdf_path.name)
            summary_sheet.append([pdf_path.name, VAL_ERROR, "", "", "", "", "", VAL_NO])

            if move_files and failed_dir:
                shutil.move(str(pdf_path), str(failed_dir / pdf_path.name))

    workbook.save(output_file)
    logging.info("Done. Output written to %s", output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract and standardize revenue data from annual-report PDFs.")
    parser.add_argument("--input-dir", required=True, help="Folder containing public annual-report PDFs.")
    parser.add_argument("--output-file", default="output/revenue_extraction_output.xlsx", help="Output Excel file path.")
    parser.add_argument("--processed-dir", default="output/processed_files", help="Folder for successfully processed PDFs.")
    parser.add_argument("--failed-dir", default="output/failed_files", help="Folder for PDFs that require review.")
    parser.add_argument("--move-files", action="store_true", help="Move processed/failed PDFs into output folders.")
    parser.add_argument("--start-page", type=int, default=4, help="First page to scan, using 1-based page numbering.")
    parser.add_argument("--max-pages", type=int, default=30, help="Maximum number of pages to scan per PDF.")
    return parser.parse_args()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()

    run_pipeline(
        input_dir=Path(args.input_dir),
        output_file=Path(args.output_file),
        processed_dir=Path(args.processed_dir),
        failed_dir=Path(args.failed_dir),
        move_files=args.move_files,
        start_page=args.start_page,
        max_pages=args.max_pages,
    )
